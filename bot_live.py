from binance.client import Client
import pandas as pd
import time
from datetime import datetime, UTC
import requests
import os
import json
import argparse
import openpyxl
import math
from db_live import init_db, insert_trade, update_strategy_status, update_commission
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from threading import Thread, Lock
from queue import Queue

# ===== ЗАГРУЗКА КОНФИГА =====
parser = argparse.ArgumentParser()
parser.add_argument("--config", required=True)
args = parser.parse_args()

with open(args.config, "r") as f:
    config = json.load(f)

BOT_NAME = config["NAME"]
load_dotenv()

# ================= НАСТРОЙКИ ОРДЕРОВ =================
MARGIN_USDT   = float(config.get("MARGIN_USDT", 10.0))   # маржа на позицию в USDT
LEVERAGE      = int(config.get("LEVERAGE", 10))           # плечо
STRATEGY_NAME = "12:4"                                     # единственная стратегия

# ================= НАСТРОЙКИ СИГНАЛОВ =================
MIN_24H_VOLUME   = config["MIN_24H_VOLUME"]
LOOKBACK_CANDLES = config["LOOKBACK_CANDLES"]
VOLUME_LOOKBACK  = config["VOLUME_LOOKBACK"]
VOL_MULT         = float(config["VOL_MULT"])
MIN_BODY_PCT     = float(config["MIN_BODY_PCT"])
COOLDOWN_BARS    = config["COOLDOWN_BARS"]
EMA_FAST         = config["EMA_FAST"]
EMA_SLOW         = config["EMA_SLOW"]
BTC_LOOKBACK     = config["BTC_LOOKBACK"]
ATR_LEN          = config["ATR_LEN"]
USE_EMA_FILTER   = config.get("USE_EMA_FILTER", True)
USE_VWAP_FILTER  = config.get("USE_VWAP_FILTER", False)
SKIP_DAYS        = [d.lower() for d in config.get("SKIP_DAYS", [])]

# Корреляция
USE_CORREL_FILTER = config.get("USE_CORREL_FILTER", False)
CORREL_MIN_BUY    = float(config.get("CORREL_MIN_BUY", 0))
CORREL_MAX_BUY    = float(config.get("CORREL_MAX_BUY", 0))
CORREL_MIN_SELL   = float(config.get("CORREL_MIN_SELL", 0))
CORREL_MAX_SELL   = float(config.get("CORREL_MAX_SELL", 0))

CHAT_ID   = os.getenv("CHAT_ID")
BOT_TOKEN = os.getenv("BOT_TOKEN")
API_KEY    = os.getenv("API_KEY_LIVE")
API_SECRET = os.getenv("API_SECRET_LIVE")

# ================= КЛИЕНТ =================
# Для тестнета:
client = Client(API_KEY, API_SECRET)
client.FUTURES_URL = "https://demo-fapi.binance.com/fapi"
# Для реального счёта — убери testnet=True:
# client = Client(API_KEY, API_SECRET)

BLACKLIST = {
    "BTCUSDT", "ETHUSDT", "BNBUSDT", "SOLUSDT",
    "XRPUSDT", "ADAUSDT", "DOGEUSDT", "LINKUSDT",
    "INTCUSDT",
}

# Стратегия 12:4 из конфига
_strat_raw = config.get("STRATEGIES_CONFIG", {})
_s124 = _strat_raw.get("12:4", {"enabled": True, "tp": 0.12, "sl": 0.04, "BUY": {}, "SELL": {}})
STRAT_CFG = {
    "tp":   float(_s124["tp"]),
    "sl":   float(_s124["sl"]),
    "BUY":  _s124.get("BUY", {}),
    "SELL": _s124.get("SELL", {}),
}
print(f"✅ Стратегия 12:4 | TP={STRAT_CFG['tp']*100}% SL={STRAT_CFG['sl']*100}% | Маржа={MARGIN_USDT}$ Плечо={LEVERAGE}x")

# ================= ФАЙЛЫ =================
TRADE_STATE_FILE   = f"trades_state_{BOT_NAME}_live.json"
EXCEL_FILE         = f"trades_{BOT_NAME}_live.xlsx"
ACTIVE_TRADES_FILE = f"active_trades_{BOT_NAME}_live.json"

TRADES_LOCK = Lock()
EXCEL_LOCK  = Lock()
_ID_LOCK    = Lock()

# ================= SYMBOL INFO CACHE =================
_symbol_info_cache = {}
_symbol_info_lock  = Lock()

def get_symbol_info(symbol):
    """Загружает и кэширует точность и минимальный размер для символа."""
    with _symbol_info_lock:
        if symbol in _symbol_info_cache:
            return _symbol_info_cache[symbol]
    try:
        info = client.futures_exchange_info()
        for s in info["symbols"]:
            if s["symbol"] == symbol:
                qty_precision   = s["quantityPrecision"]
                price_precision = s["pricePrecision"]
                min_qty = 0.0
                for f in s["filters"]:
                    if f["filterType"] == "LOT_SIZE":
                        min_qty = float(f["minQty"])
                result = {
                    "qty_precision":   qty_precision,
                    "price_precision": price_precision,
                    "min_qty":         min_qty,
                }
                with _symbol_info_lock:
                    _symbol_info_cache[symbol] = result
                return result
    except Exception as e:
        print(f"Ошибка get_symbol_info {symbol}: {e}")
    return {"qty_precision": 3, "price_precision": 4, "min_qty": 0.001}

def calc_quantity(symbol, entry_price):
    """Считает количество монет для позиции с заданной маржой и плечом."""
    info = get_symbol_info(symbol)
    notional = MARGIN_USDT * LEVERAGE          # размер позиции в USDT
    qty = notional / entry_price
    qty = round(qty, info["qty_precision"])
    qty = max(qty, info["min_qty"])
    return qty

# ================= ОРДЕРА =================
def setup_symbol(symbol, side):
    """Устанавливает плечо и изолированную маржу для символа."""
    try:
        client.futures_change_leverage(symbol=symbol, leverage=LEVERAGE)
        client.futures_change_margin_type(symbol=symbol, marginType="ISOLATED")
        print(f"⚙️ {symbol}: плечо {LEVERAGE}x, маржа ISOLATED")
    except Exception as e:
        # Ошибка -4046 = маржа уже ISOLATED — не критично
        if "-4046" not in str(e):
            print(f"⚠️ setup_symbol {symbol}: {e}")

def open_order(symbol, side, qty):
    """Открывает рыночный ордер в Hedge Mode."""
    position_side = "LONG" if side == "BUY" else "SHORT"
    try:
        order = client.futures_create_order(
            symbol=symbol,
            side=side,                        # "BUY" или "SELL"
            positionSide=position_side,       # "LONG" или "SHORT"
            type="MARKET",
            quantity=qty,
        )
        print(f"✅ Ордер открыт: {symbol} {side} {qty} | orderId={order['orderId']}")
        return order
    except Exception as e:
        print(f"❌ Ошибка открытия ордера {symbol} {side}: {e}")
        return None

def close_order(symbol, side, qty):
    position_side = "LONG" if side == "BUY" else "SHORT"
    close_side    = "SELL" if side == "BUY" else "BUY"
    try:
        order = client.futures_create_order(
            symbol=symbol,
            side=close_side,
            positionSide=position_side,
            type="MARKET",
            quantity=qty,
            # убрали reduceOnly
        )
        print(f"✅ Позиция закрыта: {symbol} {side} {qty} | orderId={order['orderId']}")
        return order
    except Exception as e:
        print(f"❌ Ошибка закрытия ордера {symbol} {side}: {e}")
        return None

# ================= TRADES =================
def load_trade_id():
    if not os.path.exists(TRADE_STATE_FILE):
        return 0
    with open(TRADE_STATE_FILE, "r") as f:
        return json.load(f).get("last_trade_id", 0)

def save_trade_id(tid):
    with open(TRADE_STATE_FILE, "w") as f:
        json.dump({"last_trade_id": tid}, f)

def save_active_trades():
    with TRADES_LOCK:
        with open(ACTIVE_TRADES_FILE, "w") as f:
            json.dump(ACTIVE_TRADES, f)

def load_active_trades():
    if not os.path.exists(ACTIVE_TRADES_FILE):
        return {}
    with open(ACTIVE_TRADES_FILE, "r") as f:
        return json.load(f)

ACTIVE_TRADES = load_active_trades()
LAST_TRADE_ID = load_trade_id()

def get_next_trade_id():
    global LAST_TRADE_ID
    with _ID_LOCK:
        LAST_TRADE_ID += 1
        save_trade_id(LAST_TRADE_ID)
        return f"LV{LAST_TRADE_ID:05d}"   # LV = Live

def check_and_close_strategies(symbol, price_high, price_low):
    """Проверяет TP/SL и закрывает позиции на бирже."""
    closed_trades = []
    with TRADES_LOCK:
        for trade_id, trade in list(ACTIVE_TRADES.items()):
            if trade["symbol"] != symbol:
                continue
            strat = trade.get("strategy")
            if not strat or strat["status"] != "OPEN":
                continue

            result = None
            if trade["side"] == "BUY":
                if price_low  <= strat["sl"]: result = "SL"
                elif price_high >= strat["tp"]: result = "TP"
            else:
                if price_high >= strat["sl"]: result = "SL"
                elif price_low  <= strat["tp"]: result = "TP"

            if result:
                # Считаем реальный PnL
                entry = trade["entry_price"]
                if trade["side"] == "BUY":
                    real_pnl = (price_high - entry) / entry * 100 if result == "TP" else (price_low - entry) / entry * 100
                else:
                    real_pnl = (entry - price_low) / entry * 100 if result == "TP" else (entry - price_high) / entry * 100
                strat["status"]     = result
                strat["close_time"] = datetime.now(UTC).strftime("%Y-%m-%d %H:%M:%S")
                update_strategy_status(trade_id, STRATEGY_NAME, result, pnl_pct=round(real_pnl, 2))

                # Закрываем реальную позицию на бирже
                qty = trade.get("qty", 0)
                if qty > 0:
                    def close_and_save_commission(sym, sid, q, tid):
                        order = close_order(sym, sid, q)
                        if order:
                            open_order_id  = trade.get("open_order_id", 0)
                            close_order_id = order.get("orderId", 0)
                            comm_open  = get_commission(sym, open_order_id)
                            comm_close = get_commission(sym, close_order_id)
                            total_comm = round(comm_open + comm_close, 6)
                            from db_live import update_commission
                            update_commission(tid, total_comm)
                            print(f"💸 Комиссия {sym}: {total_comm}")
                        Thread(target=update_trade_status_in_excel, args=(tid, result), daemon=True).start()
                    Thread(
                        target=close_and_save_commission,
                        args=(symbol, trade["side"], qty, trade_id),
                        daemon=True
                    ).start()

                send_telegram(
                    f"{'✅' if result=='TP' else '❌'} {result} | {BOT_NAME}\n"
                    f"{symbol} {trade['side']} | {STRATEGY_NAME}\n"
                    f"Вход: {trade['entry_price']} | Qty: {qty}"
                )

            if strat["status"] != "OPEN":
                closed_trades.append(trade_id)

        for tid in closed_trades:
            del ACTIVE_TRADES[tid]

    if closed_trades:
        save_active_trades()
    return closed_trades

def get_commission(symbol, order_id):
    """Получает комиссию за конкретный ордер."""
    try:
        trades = client.futures_account_trades(symbol=symbol, limit=20)
        total = 0.0
        for t in trades:
            if str(t.get("orderId")) == str(order_id):
                total += abs(float(t.get("commission", 0)))
        return round(total, 6)
    except Exception as e:
        print(f"Ошибка комиссии {symbol}: {e}")
        return 0.0

# ================= TELEGRAM =================
def send_telegram(message: str):
    if not BOT_TOKEN or not CHAT_ID:
        return
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    try:
        requests.post(url, data={"chat_id": CHAT_ID, "text": message}, timeout=10)
    except Exception as e:
        print(f"Ошибка Telegram: {e}")

# ================= EXCEL =================
def write_trade_to_excel(trade_id, symbol, side, entry_price, qty, tp, sl, corr_text, natr, vol24):
    with EXCEL_LOCK:
        if not os.path.exists(EXCEL_FILE):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "live"
            ws.append(["Trade_id","Дата","Время","Символ","Сторона","Вход","Qty","TP","SL","Маржа","Плечо","Corr","NATR","Vol24h","Статус"])
            wb.save(EXCEL_FILE)

        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        dt = datetime.now()
        ws.append([
            trade_id,
            dt.strftime("%d.%m.%Y"),
            dt.strftime("%H:%M:%S"),
            symbol, side,
            entry_price, qty, tp, sl,
            MARGIN_USDT, LEVERAGE,
            corr_text, natr, round(vol24/1_000_000, 1),
            "OPEN"
        ])
        wb.save(EXCEL_FILE)

def update_trade_status_in_excel(trade_id, status):
    with EXCEL_LOCK:
        if not os.path.exists(EXCEL_FILE):
            return
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row, 1).value) == trade_id:
                ws.cell(row, 15).value = status  # колонка "Статус"
                break
        wb.save(EXCEL_FILE)

# ================= INDICATORS =================
def calculate_session_vwap(df):
    df = df.copy()
    df["date"] = pd.to_datetime(df["open_time"], unit="ms").dt.date
    tp = (df["high"] + df["low"] + df["close"]) / 3
    df["tpv"] = tp * df["volume"]
    df["cum_tpv"] = df.groupby("date")["tpv"].cumsum()
    df["cum_vol"] = df.groupby("date")["volume"].cumsum()
    return df["cum_tpv"] / df["cum_vol"]

def calculate_atr(df, period):
    hl = df["high"] - df["low"]
    hc = (df["high"] - df["close"].shift()).abs()
    lc = (df["low"]  - df["close"].shift()).abs()
    tr = pd.concat([hl, hc, lc], axis=1).max(axis=1)
    return tr.rolling(period).mean()

def get_liquid_futures_symbols():
    tickers = client._request_futures_api(method="get", path="ticker/24hr")
    symbols = []
    for t in tickers:
        sym = t["symbol"]
        if not sym.endswith("USDT") or sym in BLACKLIST:
            continue
        if "_" in sym:          
            continue
        if float(t["quoteVolume"]) < MIN_24H_VOLUME:
            continue
        symbols.append(sym)
    return symbols

def get_symbols_with_open_trades():
    with TRADES_LOCK:
        return {
            trade["symbol"] for trade in ACTIVE_TRADES.values()
            if trade.get("strategy", {}).get("status") == "OPEN"
        }

def get_btc_returns():
    try:
        klines = client.futures_klines(symbol="BTCUSDT", interval=Client.KLINE_INTERVAL_1HOUR, limit=BTC_LOOKBACK)
        df = pd.DataFrame(klines, columns=[
            "open_time","open","high","low","close","volume",
            "close_time","quote_volume","trades","taker_buy_base","taker_buy_quote","ignore"
        ])
        df["close"] = df["close"].astype(float)
        return df["close"].pct_change()
    except Exception as e:
        print(f"Ошибка BTC returns: {e}")
        return None

def apply_range_filter(value, min_val, max_val):
    if value is None: return False
    if min_val != 0 and value < min_val: return False
    if max_val != 0 and value > max_val: return False
    return True

def check_strategy_filters(side, natr, delta_pct, corr):
    """Проверяет фильтры стратегии 12:4 для заданной стороны."""
    f = STRAT_CFG.get(side, {})
    if not f:
        return True

    today = datetime.now(UTC).strftime("%A").lower()
    if today in [d.lower() for d in f.get("SKIP_DAYS", [])]:
        return False

    if f.get("USE_NATR_FILTER"):
        if not apply_range_filter(natr, f.get("NATR_MIN", 0), f.get("NATR_MAX", 0)):
            return False

    if f.get("USE_DELTA_FILTER"):
        val = delta_pct if side == "BUY" else -delta_pct
        if not apply_range_filter(val, f.get("DELTA_MIN", 0), f.get("DELTA_MAX", 0)):
            return False

    if f.get("USE_CORREL_FILTER") and corr is not None:
        try:
            if not apply_range_filter(float(corr), f.get("CORREL_MIN", 0), f.get("CORREL_MAX", 0)):
                return False
        except (ValueError, TypeError):
            pass

    return True

def calculate_delta(last_row):
    rng = last_row["high"] - last_row["low"]
    if rng == 0: return 0.0
    buy_vol  = last_row["volume"] * (last_row["close"] - last_row["low"]) / rng
    sell_vol = last_row["volume"] - buy_vol
    return round((buy_vol - sell_vol) / last_row["volume"] * 100, 2)

def check_volume_signal(symbol):
    klines = client.futures_klines(
        symbol=symbol, interval=Client.KLINE_INTERVAL_1HOUR, limit=LOOKBACK_CANDLES
    )
    df = pd.DataFrame(klines, columns=[
        "open_time","open","high","low","close",
        "volume","close_time","quote_volume",
        "trades","taker_buy_base","taker_buy_quote","ignore"
    ])
    for c in ["open","high","low","close","volume"]:
        df[c] = df[c].astype(float)

    df["ema20"]  = df["close"].ewm(span=EMA_FAST,  adjust=False).mean()
    df["ema200"] = df["close"].ewm(span=EMA_SLOW,  adjust=False).mean()
    df["atr"]    = calculate_atr(df, ATR_LEN)
    df["natr"]   = (df["atr"] / df["close"]) * 100
    df["vwap"]   = calculate_session_vwap(df)
    df["quote_volume"] = df["close"] * df["volume"]

    avg_vol = df["quote_volume"].iloc[-(VOLUME_LOOKBACK + 2):-2].mean()
    last    = df.iloc[-2]

    volume_spike = last["quote_volume"] >= avg_vol * VOL_MULT
    body      = abs(last["close"] - last["open"])
    rng       = last["high"]  - last["low"]
    body_pct  = 0 if rng == 0 else body / rng * 100
    bull      = last["close"] > last["open"]
    bear      = last["close"] < last["open"]
    strong    = body_pct >= MIN_BODY_PCT

    bull_trend  = last["ema20"] > last["ema200"]
    bear_trend  = last["ema20"] < last["ema200"]
    ema_bull_ok = bull_trend if USE_EMA_FILTER else True
    ema_bear_ok = bear_trend if USE_EMA_FILTER else True

    below_vwap = (last["close"] < last["vwap"]) if USE_VWAP_FILTER else True
    above_vwap = (last["close"] > last["vwap"]) if USE_VWAP_FILTER else True

    recent_spike = False
    if COOLDOWN_BARS > 0:
        recent = df.iloc[-(COOLDOWN_BARS + 2):-2]
        recent_spike = (recent["quote_volume"] >= avg_vol * VOL_MULT).any()

    delta_pct = calculate_delta(last)

    signals = []
    if volume_spike and bull and strong and ema_bull_ok and below_vwap and not recent_spike:
        signals.append("BUY_TREND")
    if volume_spike and bear and strong and ema_bear_ok and above_vwap and not recent_spike:
        signals.append("SELL_TREND")

    if not signals:
        return None

    ticker_24h = client.futures_ticker(symbol=symbol)
    volume_24h = float(ticker_24h["quoteVolume"])

    return {
        "symbol":    symbol,
        "signals":   signals,
        "close":     last["close"],
        "natr":      round(last["natr"], 3),
        "volText":   f"x{last['quote_volume']/avg_vol:.2f}",
        "volume_24h": volume_24h,
        "delta_pct": delta_pct,
    }

# ================= МОНИТОРИНГ TP/SL =================
_monitored_symbols = set()
_monitor_lock      = Lock()

def start_price_monitor(symbol):
    with _monitor_lock:
        if symbol in _monitored_symbols:
            return
        _monitored_symbols.add(symbol)

    def monitor():
        print(f"👁️ Мониторинг: {symbol}")
        while True:
            try:
                with TRADES_LOCK:
                    has_open = any(
                        t["symbol"] == symbol and t.get("strategy", {}).get("status") == "OPEN"
                        for t in ACTIVE_TRADES.values()
                    )
                if not has_open:
                    with _monitor_lock:
                        _monitored_symbols.discard(symbol)
                    print(f"👁️ Мониторинг остановлен: {symbol}")
                    break
                ticker = client.futures_symbol_ticker(symbol=symbol)
                # Если вернулся список — берём первый элемент
                if isinstance(ticker, list):
                    ticker = ticker[0]
                price = float(ticker["price"])
                if price > 0:
                    check_and_close_strategies(symbol, price, price)
            except Exception as e:
                print(f"Ошибка мониторинга {symbol}: {e}")
            time.sleep(1)

    Thread(target=monitor, daemon=True).start()

def sync_active_trades_with_db():
    import sqlite3
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "trades_live.db")
    while True:
        time.sleep(30)
        try:
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            with TRADES_LOCK:
                for trade_id, trade in list(ACTIVE_TRADES.items()):
                    rows = conn.execute(
                        "SELECT strategy, status FROM trade_strategies WHERE trade_id = ?",
                        (trade_id,)
                    ).fetchall()
                    for row in rows:
                        if row["strategy"] == STRATEGY_NAME and trade.get("strategy", {}).get("status") == "OPEN":
                            if row["status"] not in ("OPEN",):
                                trade["strategy"]["status"] = row["status"]
                                print(f"🔄 Синхронизировано: {trade_id} → {row['status']}")
                    if trade.get("strategy", {}).get("status") != "OPEN":
                        del ACTIVE_TRADES[trade_id]
            conn.close()
            save_active_trades()
        except Exception as e:
            print(f"Ошибка sync: {e}")

_need_restart = [False]

# ================= MAIN =================
def main():
    global _need_restart
    init_db()

    # Стартуем мониторинг для уже открытых позиций
    for trade in ACTIVE_TRADES.values():
        if trade.get("strategy", {}).get("status") == "OPEN":
            start_price_monitor(trade["symbol"])

    Thread(target=sync_active_trades_with_db, daemon=True).start()

    symbols = get_liquid_futures_symbols()
    print(f"✅ Ликвидных токенов: {len(symbols)}")

    last_signal_time = {}
    cooldown_seconds = COOLDOWN_BARS * 3600

    def update_symbols_periodically():
        nonlocal symbols
        while True:
            time.sleep(3600)
            try:
                liquid    = get_liquid_futures_symbols()
                open_syms = get_symbols_with_open_trades()
                symbols   = list(set(liquid) | open_syms)
                print(f"♻️ Обновление: {len(liquid)} ликвидных + {len(open_syms)} открытых")
            except Exception as e:
                print(f"Ошибка обновления символов: {e}")

    Thread(target=update_symbols_periodically, daemon=True).start()

    task_queue = Queue()

    def process_signal(msg):
        today = datetime.now(UTC).strftime("%A").lower()
        if today in SKIP_DAYS:
            return
        try:
            if msg.get("e") == "error":
                err_msg = msg.get("m", "")
                print(f"🔴 WS ошибка: {msg}")
                if "reset" in err_msg.lower() or "closed" in err_msg.lower():
                    send_telegram(f"🔴 {BOT_NAME} WS ошибка: {err_msg}")
                    _need_restart[0] = True
                return

            if "data" not in msg or "k" not in msg["data"]:
                return
            candle = msg["data"]["k"]
            symbol = candle["s"]
            if not candle["x"]:
                return

            # Проверка TP/SL по свече
            check_and_close_strategies(symbol, float(candle["h"]), float(candle["l"]))

            if symbol not in symbols:
                return

            now = time.time()
            if now - last_signal_time.get(symbol, 0) < cooldown_seconds:
                return

            res = check_volume_signal(symbol)
            if not res:
                return

            last_signal_time[symbol] = now

            entry_price = res["close"]
            side = "BUY" if any("BUY" in s for s in res["signals"]) else "SELL"

            # Корреляция BTC
            corr_text = "N/A"
            try:
                btc_ret = get_btc_returns()
                if btc_ret is not None:
                    klines_sym = client.futures_klines(
                        symbol=symbol, interval=Client.KLINE_INTERVAL_1HOUR, limit=BTC_LOOKBACK
                    )
                    df_sym = pd.DataFrame(klines_sym, columns=[
                        "open_time","open","high","low","close","volume",
                        "close_time","quote_volume","trades","taker_buy_base","taker_buy_quote","ignore"
                    ])
                    df_sym["close"] = df_sym["close"].astype(float)
                    sym_ret = df_sym["close"].pct_change()
                    corr = btc_ret[-len(sym_ret):].corr(sym_ret)
                    corr_text = round(float(corr), 2) if corr is not None else "N/A"
            except Exception as e:
                print(f"Ошибка корреляции {symbol}: {e}")

            # Глобальный фильтр корреляции
            if USE_CORREL_FILTER and isinstance(corr_text, float):
                mn = CORREL_MIN_BUY if side == "BUY" else CORREL_MIN_SELL
                mx = CORREL_MAX_BUY if side == "BUY" else CORREL_MAX_SELL
                if not apply_range_filter(corr_text, mn, mx):
                    print(f"⏭️ {symbol} пропущен — корреляция {corr_text}")
                    return

            # Фильтры стратегии 12:4
            if not check_strategy_filters(side, res["natr"], res["delta_pct"], corr_text):
                print(f"⏭️ {symbol} {side} — не прошёл фильтры 12:4")
                return

            # Проверяем нет ли уже открытой позиции на этот символ+сторону
            with TRADES_LOCK:
                already_open = any(
                    t["symbol"] == symbol and t["side"] == side
                    and t.get("strategy", {}).get("status") == "OPEN"
                    for t in ACTIVE_TRADES.values()
                )
            if already_open:
                print(f"⏭️ {symbol} {side} — уже есть открытая позиция")
                return

            # Устанавливаем плечо и маржу
            setup_symbol(symbol, side)

            # Считаем количество
            qty = calc_quantity(symbol, entry_price)
            if qty <= 0:
                print(f"❌ {symbol} qty=0, пропускаем")
                return

            # Открываем ордер на бирже
            order = open_order(symbol, side, qty)
            if not order:
                return

            # Реальная цена входа из ордера (если market — close ≈ entry)
            real_entry = float(order.get("avgPrice") or entry_price)
            if real_entry == 0:
                real_entry = entry_price

            # Уровни TP/SL
            if side == "BUY":
                tp = real_entry * (1 + STRAT_CFG["tp"])
                sl = real_entry * (1 - STRAT_CFG["sl"])
            else:
                tp = real_entry * (1 - STRAT_CFG["tp"])
                sl = real_entry * (1 + STRAT_CFG["sl"])

            trade_id = get_next_trade_id()

            with TRADES_LOCK:
                ACTIVE_TRADES[trade_id] = {
                    "open_order_id": order.get("orderId", 0),
                    "symbol":      symbol,
                    "side":        side,
                    "entry_price": real_entry,
                    "qty":         qty,
                    "open_time":   datetime.now(UTC).strftime("%Y-%m-%d %H:%M:%S"),
                    "strategy": {
                        "tp":     tp,
                        "sl":     sl,
                        "status": "OPEN",
                    }
                }
            save_active_trades()

            # Запускаем мониторинг
            start_price_monitor(symbol)

            # Пишем в Excel и БД
            write_trade_to_excel(
                trade_id, symbol, side, real_entry, qty,
                round(tp, 6), round(sl, 6),
                corr_text, res["natr"], res["volume_24h"]
            )
            insert_trade(trade_id, BOT_NAME, {
                "symbol":      symbol,
                "side":        side,
                "entry_price": real_entry,
                "signals":     res["signals"],
                "strategies":  {STRATEGY_NAME: {"tp": tp, "sl": sl, "status": "OPEN"}},
                "natr":        res["natr"],
                "swing_num":   0,
                "delta_pct":   res["delta_pct"],
                "relvol":      "N/A",
                "atr_ratio_50": 0, "atr_ratio_75": 0, "atr_ratio_100": 0,
                "expansion_3": 0,  "expansion_5":  0,
            }, vol_text=res["volText"], vol_24h=res["volume_24h"]/1_000_000, corr_text=corr_text)

            # Telegram
            vol24 = res["volume_24h"] / 1_000_000
            send_telegram(
                f"🤖 {BOT_NAME} | 12:4\n"
                f"{'🟢' if side=='BUY' else '🔴'} {symbol} {side}\n"
                f"Вход: {real_entry:.6f} | Qty: {qty}\n"
                f"TP: {tp:.6f} | SL: {sl:.6f}\n"
                f"Маржа: {MARGIN_USDT}$ | Плечо: {LEVERAGE}x\n"
                f"Vol: {res['volText']} | {vol24:.1f}M USDT\n"
                f"NATR: {res['natr']}% | Corr: {corr_text}"
            )
            print(f"🔥 {symbol} {side} | entry={real_entry} tp={tp:.4f} sl={sl:.4f} qty={qty}")

        except Exception as e:
            print(f"Ошибка process_signal: {e}")
            import traceback; traceback.print_exc()

    def handle_kline(msg):
        task_queue.put(msg)

    def worker():
        while True:
            msg = task_queue.get()
            process_signal(msg)
            task_queue.task_done()

    Thread(target=worker, daemon=True).start()

    # ===== REST polling вместо WebSocket =====
    last_candle_time = {}
    print("🟢 REST polling запущен")
    send_telegram(f"🟢 {BOT_NAME} Live REST polling запущен")

    while True:
        try:
            now = datetime.now(UTC)
            if now.minute == 0 and now.second < 120:
                open_syms   = get_symbols_with_open_trades()
                all_symbols = list(set(symbols) | open_syms)
                print(f"🔄 Polling {len(all_symbols)} символов...")

                for symbol in all_symbols:
                    try:
                        klines = client.futures_klines(
                            symbol=symbol,
                            interval=Client.KLINE_INTERVAL_1HOUR,
                            limit=2
                        )
                        candle_time = klines[-2][0]

                        if last_candle_time.get(symbol) == candle_time:
                            continue
                        last_candle_time[symbol] = candle_time

                        high = float(klines[-2][2])
                        low  = float(klines[-2][3])
                        task_queue.put({
                            "data": {
                                "k": {
                                    "s": symbol,
                                    "x": True,
                                    "h": str(high),
                                    "l": str(low),
                                }
                            }
                        })
                    except Exception as e:
                        print(f"Ошибка poll {symbol}: {e}")
                    time.sleep(0.05)

        except Exception as e:
            print(f"Ошибка polling loop: {e}")

        time.sleep(10)


if __name__ == "__main__":
    main()
