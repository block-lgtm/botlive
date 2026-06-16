"""
Live Trading API — порт 8766
Запуск: uvicorn api_live:app --host 0.0.0.0 --port 8766
"""

from fastapi import FastAPI, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
import asyncio
import json
import os
import secrets
from fastapi import Depends, HTTPException
from fastapi.responses import HTMLResponse, FileResponse
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from datetime import datetime, timezone
from dotenv import load_dotenv
from db_live import (
    get_open_trades, get_closed_trades, get_stats,
    get_daily_stats, manual_close_strategy, get_symbol_stats, get_weekday_stats,
    get_total_commission
)

security = HTTPBasic()

load_dotenv()
API_KEY    = os.getenv("API_KEY_LIVE")
API_SECRET = os.getenv("API_SECRET_LIVE")

from binance.client import Client as BinanceClient
_binance_client = BinanceClient(API_KEY, API_SECRET)

app = FastAPI(title="Live Trading Dashboard")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

STRATEGY_NAME = "12:4"

def check_auth(credentials: HTTPBasicCredentials = Depends(security)):
    pwd = os.getenv("DASHBOARD_PASSWORD", "changeme")
    correct_password = secrets.compare_digest(credentials.password, pwd)
    correct_username = secrets.compare_digest(credentials.username, "admin")
    if not (correct_username and correct_password):
        raise HTTPException(status_code=401, headers={"WWW-Authenticate": "Basic"})
    return credentials

@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(credentials: HTTPBasicCredentials = Depends(check_auth)):
    with open("/root/bot/botlive/dashboard_live.html", "r", encoding="utf-8") as f:
        return f.read()

def group_trades_by_id(rows):
    trades = {}
    for row in rows:
        tid = row["id"]
        if tid not in trades:
            trades[tid] = {
                "id":          row["id"],
                "bot_name":    row["bot_name"],
                "symbol":      row["symbol"],
                "side":        row["side"],
                "entry_price": row["entry_price"],
                "qty":         row.get("qty"),
                "open_time":   row["open_time"],
                "close_time":  row.get("close_time"),
                "natr":        row["natr"],
                "vol_text":    row["vol_text"],
                "vol_24h":     row["vol_24h"],
                "corr_btc":    row["corr_btc"],
                "signals":     row["signals"],
                "delta_pct":   row["delta_pct"],
                "commission": row.get("commission", 0),
                "strategies":  {},
            }
        trades[tid]["strategies"][row["strategy"]] = {
            "tp":         row["tp"],
            "sl":         row["sl"],
            "status":     row["strat_status"],
            "close_time": row.get("strat_close"),
             "pnl_pct":    row.get("pnl_pct"),
        }
    return list(trades.values())


def enrich_open_trade(trade):
    try:
        open_dt = datetime.strptime(trade["open_time"], "%Y-%m-%d %H:%M:%S")
        open_dt = open_dt.replace(tzinfo=timezone.utc)
        delta = datetime.now(timezone.utc) - open_dt
        hours   = int(delta.total_seconds() // 3600)
        minutes = int((delta.total_seconds() % 3600) // 60)
        trade["duration"] = f"{hours}ч {minutes}м"
    except Exception:
        trade["duration"] = "—"
    return trade


def get_balance():
    """Получает баланс с Binance Futures."""
    try:
        from binance.client import Client
        c = Client(API_KEY, API_SECRET)
        account = c.futures_account_balance()
        total = 0.0
        for b in account:
            if b["asset"] in ("USDT", "USDC", "BNFCR"):
                val = float(b["balance"])
                if val > 0:
                    total += val
        return round(total, 2) if total > 0 else None
    except Exception as e:
        print(f"Ошибка баланса: {e}")
    return None


# ── REST ──

@app.get("/trades/open")
def open_trades():
    rows   = get_open_trades()
    trades = group_trades_by_id(rows)
    trades = [enrich_open_trade(t) for t in trades]
    return {"trades": trades, "count": len(trades)}


@app.get("/trades/closed")
def closed_trades(limit: int = 2000):
    rows   = get_closed_trades(limit=limit)
    trades = group_trades_by_id(rows)
    return {"trades": trades, "count": len(trades)}


@app.get("/stats")
def stats():
    return get_stats()


@app.get("/stats/daily")
def daily_stats(date_from: str = None, date_to: str = None, strategies: str = None, side: str = None):
    strats = strategies.split(",") if strategies else None
    return get_daily_stats(date_from=date_from, date_to=date_to, strategies=strats, side=side)

@app.get("/stats/symbols")
def symbol_stats(date_from: str = None, date_to: str = None, strategies: str = None, side: str = None):
    strats = strategies.split(",") if strategies else None
    return get_symbol_stats(date_from=date_from, date_to=date_to, strategies=strats, side=side)

@app.get("/stats/weekdays")
def weekday_stats(date_from: str = None, date_to: str = None, strategies: str = None, side: str = None):
    strats = strategies.split(",") if strategies else None
    return get_weekday_stats(date_from=date_from, date_to=date_to, strategies=strats, side=side)

@app.get("/stats/commission")
def commission_stats():
    return {"total_commission": get_total_commission()}

def send_telegram(message: str):
    bot_token = os.getenv("BOT_TOKEN")
    chat_id = os.getenv("CHAT_ID")
    if not bot_token or not chat_id:
        return
    try:
        import requests
        requests.post(f"https://api.telegram.org/bot{bot_token}/sendMessage",
                     data={"chat_id": chat_id, "text": message}, timeout=10)
    except Exception as e:
        print(f"Ошибка Telegram: {e}")


@app.post("/trades/{trade_id}/close/{strategy}")
def close_strategy(trade_id: str, strategy: str, price: float = 0):
    # Сначала закрываем на бирже
    try:
        rows = get_open_trades()
        trade = next((t for t in rows if t["id"] == trade_id), None)
        if trade:
            symbol = trade["symbol"]
            side = trade["side"]
            qty = float(trade.get("qty") or 0)
            print(f"🔍 Ручное закрытие: {symbol} {side} qty={qty}")
            if qty > 0:
                position_side = "LONG" if side == "BUY" else "SHORT"
                close_side = "SELL" if side == "BUY" else "BUY"
                _binance_client.futures_create_order(
                    symbol=symbol,
                    side=close_side,
                    positionSide=position_side,
                    type="MARKET",
                    quantity=qty,
                )
                print(f"✅ Ручное закрытие на бирже: {symbol} {side} {qty}")
        else:
            print(f"❌ Сделка {trade_id} не найдена в открытых")
    except Exception as e:
        print(f"❌ Ошибка ручного закрытия на бирже: {e}")
        import traceback; traceback.print_exc()

    result = manual_close_strategy(trade_id, strategy, price)
    if "error" not in result:
        pnl = result.get("pnl_pct", 0)
        status = result.get("status", "—")
        send_telegram(
            f"🖐 Ручное закрытие\n"
            f"{trade_id} | {strategy}\n"
            f"Результат: {status} | PnL: {'+' if pnl>=0 else ''}{pnl:.2f}%"
        )
    if "error" not in result:
        # Обновляем Excel
        try:
            import openpyxl
            excel_file = f"/root/bot/botlive/trades_LIVE12_4_live.xlsx"
            if os.path.exists(excel_file):
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active
                for row in range(2, ws.max_row + 1):
                    if str(ws.cell(row, 1).value) == trade_id:
                        ws.cell(row, 15).value = result["status"]         # колонка O — статус
                        ws.cell(row, 16).value = round(result["pnl_pct"], 2)  # колонка P — результат %
                        break
                wb.save(excel_file)
        except Exception as e:
            print(f"Ошибка обновления Excel: {e}")
    return result

@app.get("/balance")
def balance():
    return {"balance": get_balance()}


@app.get("/health")
def health():
    return {"status": "ok", "time": datetime.now(timezone.utc).isoformat()}


# ── WebSocket ──

class ConnectionManager:
    def __init__(self):
        self.active: list[WebSocket] = []

    async def connect(self, ws: WebSocket):
        await ws.accept()
        self.active.append(ws)

    def disconnect(self, ws: WebSocket):
        if ws in self.active:
            self.active.remove(ws)

    async def broadcast(self, data: dict):
        msg = json.dumps(data)
        dead = []
        for ws in self.active:
            try:
                await ws.send_text(msg)
            except Exception:
                dead.append(ws)
        for ws in dead:
            self.disconnect(ws)


manager = ConnectionManager()


@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await manager.connect(websocket)
    try:
        balance = get_balance()
        await websocket.send_text(json.dumps({
            "type":    "init",
            "open":    group_trades_by_id(get_open_trades()),
            "stats":   get_stats(),
            "balance": balance,
        }))
        while True:
            try:
                data = await asyncio.wait_for(websocket.receive_text(), timeout=30)
                if data == "ping":
                    await websocket.send_text(json.dumps({"type": "pong"}))
            except asyncio.TimeoutError:
                await websocket.send_text(json.dumps({"type": "heartbeat"}))
    except WebSocketDisconnect:
        manager.disconnect(websocket)


@app.on_event("startup")
async def start_background_push():
    asyncio.create_task(push_updates())


async def push_updates():
    counter = 0
    while True:
        await asyncio.sleep(5)
        if not manager.active:
            continue
        try:
            counter += 1
            rows   = get_open_trades()
            trades = group_trades_by_id(rows)
            trades = [enrich_open_trade(t) for t in trades]
            # Добавляем текущую цену
            for t in trades:
                try:
                    ticker = _binance_client.futures_mark_price(symbol=t["symbol"])
                    if isinstance(ticker, list): ticker = ticker[0]
                    t["current_price"] = float(ticker["markPrice"])
                except:
                    t["current_price"] = None
            # Баланс обновляем раз в минуту чтобы не спамить API
            balance = get_balance() if counter % 12 == 0 else None
            payload = {
                "type":  "update",
                "open":  trades,
                "stats": get_stats(),
            }
            if balance is not None:
                payload["balance"] = balance
            await manager.broadcast(payload)
        except Exception as e:
            print(f"Ошибка push_updates: {e}")
